# services/user_service.py
from io import BytesIO
from typing import List, Dict, Any
import openpyxl
from firebase_admin import auth, exceptions
from core.firebase_db import db


def parse_excel_data(file_content: bytes, user_type: str) -> List[Dict[str, Any]]:
    """
    Parses the byte content of an Excel file into a list of user dictionaries.
    """
    try:
        workbook = openpyxl.load_workbook(BytesIO(file_content))
        sheet = workbook.active

        headers = [cell.value.lower().strip() for cell in sheet[1]]

        if user_type == "lecturer":
            required_headers = ["lecturer_id", "ic", "name", "email", "faculty"]
        elif user_type == "student":
            required_headers = [
                "student_id",
                "ic",
                "name",
                "email",
                "major",
                "year",
                "semester",
            ]
        else:
            raise ValueError(f"Invalid user type: {user_type}")

        if not all(h in headers for h in required_headers):
            missing = [h for h in required_headers if h not in headers]
            raise ValueError(f"Missing required columns: {', '.join(missing)}")

        users = []
        for row_index in range(2, sheet.max_row + 1):
            user_data = {}
            row_is_empty = True

            for col_index, header in enumerate(headers):
                value = sheet.cell(row=row_index, column=col_index + 1).value
                if value is not None:
                    row_is_empty = False

                if header in ["id", "email", "ic", "student_id", "lecturer_id"]:
                    user_data[header] = str(value).strip() if value else None
                elif header in ["year", "semester"]:
                    user_data[header] = int(value) if value else None
                else:
                    user_data[header] = value

            if not row_is_empty:
                # Basic validation
                if all(user_data.get(h) for h in required_headers):
                    final_user_data = {k: user_data.get(k) for k in required_headers}
                    users.append(final_user_data)

        if not users:
            raise ValueError("No valid user records found in the Excel file.")

        return users

    except Exception as e:
        raise Exception(f"Failed to process Excel file: {str(e)}")


def bulk_create_users(users: List[Dict[str, Any]], user_type: str) -> Dict[str, int]:
    """
    Creates or Updates users.
    FORCE updates password to IC number every time.
    """
    stats = {"total": len(users), "created": 0, "updated": 0, "failed": 0, "errors": []}
    firestore_batch = db.batch()

    for user in users:
        email = user["email"]
        ic_number = str(user["ic"])

        if user_type == "lecturer":
            user_id = str(user["lecturer_id"])
            id_field = "lecturer_id"
        else:
            user_id = str(user["student_id"])
            id_field = "student_id"

        # 1. Firebase Authentication (Create or Update)
        try:
            try:
                # Try to create
                auth.create_user(
                    uid=user_id,
                    email=email,
                    password=ic_number,  # Set Password = IC
                    display_name=user["name"],
                )
                stats["created"] += 1
            except exceptions.AlreadyExistsError:
                # If exists, FORCE UPDATE the password
                auth.update_user(
                    uid=user_id,
                    password=ic_number,  # RESET Password to IC
                    email=email,
                    display_name=user["name"],
                )
                stats["updated"] += 1

        except Exception as e:
            stats["failed"] += 1
            stats["errors"].append(f"Auth Error ({email}): {str(e)}")
            continue  # Skip Firestore update if Auth failed

        # 2. Firestore Profile
        profile_data = {
            "uid": user_id,
            "email": email,
            "name": user["name"],
            "role": user_type,
            "ic": ic_number,
            id_field: user_id,
        }

        if user_type == "student":
            profile_data["major"] = user["major"]
            profile_data["year"] = user.get("year")
            profile_data["semester"] = user.get("semester")
        elif user_type == "lecturer":
            profile_data["faculty"] = user.get("faculty")

        user_doc_ref = db.collection("users").document(user_id)
        firestore_batch.set(user_doc_ref, profile_data, merge=True)

    # Commit Firestore changes
    try:
        firestore_batch.commit()
    except Exception as e:
        stats["errors"].append(f"Firestore Batch Error: {str(e)}")

    return stats


def get_user_profile(user_id: str) -> Dict[str, Any] | None:
    try:
        doc = db.collection("users").document(user_id).get()
        if doc.exists:
            data = doc.to_dict()
            return data
        return None
    except Exception:
        return None
