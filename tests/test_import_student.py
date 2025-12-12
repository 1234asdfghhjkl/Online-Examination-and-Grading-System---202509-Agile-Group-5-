# tests/test_import_student.py
import unittest
from unittest.mock import patch, MagicMock
from io import BytesIO
import openpyxl
from firebase_admin import exceptions

# Import the functions to test
from services.user_service import parse_excel_data, bulk_create_users


class TestParseExcelData(unittest.TestCase):
    """Test cases for parse_excel_data function"""

    def create_excel_file(self, headers, data_rows):
        """Helper function to create an in-memory Excel file"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)

        # Write data rows
        for row_idx, row_data in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        # Save to BytesIO
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.read()

    # ==================== POSITIVE TEST CASES ====================

    def test_parse_valid_student_excel(self):
        """Test parsing a valid student Excel file with all required fields"""
        headers = ["student_id", "ic", "name", "email", "major", "year", "semester"]
        data_rows = [
            ["S001", "990101010101", "John Doe", "john@example.com", "Computer Science", 2, 1],
            ["S002", "980202020202", "Jane Smith", "jane@example.com", "Engineering", 3, 2],
        ]

        excel_content = self.create_excel_file(headers, data_rows)
        result = parse_excel_data(excel_content, "student")

        self.assertEqual(len(result), 2)
        self.assertEqual(result[0]["student_id"], "S001")
        self.assertEqual(result[0]["name"], "John Doe")
        self.assertEqual(result[0]["year"], 2)
        self.assertEqual(result[1]["student_id"], "S002")

    def test_parse_student_excel_with_extra_columns(self):
        """Test parsing Excel with extra columns (should be ignored)"""
        headers = ["student_id", "ic", "name", "email", "major", "year", "semester", "extra_column"]
        data_rows = [
            ["S001", "990101010101", "John Doe", "john@example.com", "CS", 2, 1, "ignore_me"],
        ]

        excel_content = self.create_excel_file(headers, data_rows)
        result = parse_excel_data(excel_content, "student")

        self.assertEqual(len(result), 1)
        self.assertNotIn("extra_column", result[0])

    def test_parse_student_excel_mixed_case_headers(self):
        """Test parsing Excel with mixed case headers (should work with .lower())"""
        headers = ["Student_ID", "IC", "Name", "EMAIL", "major", "Year", "Semester"]
        data_rows = [
            ["S001", "990101010101", "John Doe", "john@example.com", "CS", 2, 1],
        ]

        excel_content = self.create_excel_file(headers, data_rows)
        result = parse_excel_data(excel_content, "student")

        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["student_id"], "S001")

    def test_parse_student_excel_skip_empty_rows(self):
        """Test that empty rows are skipped"""
        headers = ["student_id", "ic", "name", "email", "major", "year", "semester"]
        data_rows = [
            ["S001", "990101010101", "John Doe", "john@example.com", "CS", 2, 1],
            [None, None, None, None, None, None, None],  # Empty row
            ["S002", "980202020202", "Jane Smith", "jane@example.com", "ENG", 3, 2],
        ]

        excel_content = self.create_excel_file(headers, data_rows)
        result = parse_excel_data(excel_content, "student")

        self.assertEqual(len(result), 2)
        self.assertEqual(result[0]["student_id"], "S001")
        self.assertEqual(result[1]["student_id"], "S002")

    # ==================== NEGATIVE TEST CASES ====================

    def test_parse_student_excel_missing_required_column(self):
        """Test parsing Excel missing a required column (should raise Exception)"""
        headers = ["student_id", "ic", "name", "email", "major", "year"]  # Missing 'semester'
        data_rows = [
            ["S001", "990101010101", "John Doe", "john@example.com", "CS", 2],
        ]

        excel_content = self.create_excel_file(headers, data_rows)

        with self.assertRaises(Exception) as context:
            parse_excel_data(excel_content, "student")

        self.assertIn("Failed to process Excel file", str(context.exception))
        self.assertIn("semester", str(context.exception))

    def test_parse_student_excel_no_valid_records(self):
        """Test parsing Excel with no valid records (should raise Exception)"""
        headers = ["student_id", "ic", "name", "email", "major", "year", "semester"]
        data_rows = [
            [None, None, None, None, None, None, None],  # All empty
            ["", "", "", "", "", "", ""],  # All blank strings
        ]

        excel_content = self.create_excel_file(headers, data_rows)

        with self.assertRaises(Exception) as context:
            parse_excel_data(excel_content, "student")

        self.assertIn("No valid user records found", str(context.exception))

    def test_parse_student_excel_incomplete_data(self):
        """Test parsing Excel with incomplete data in rows (should skip incomplete rows)"""
        headers = ["student_id", "ic", "name", "email", "major", "year", "semester"]
        data_rows = [
            ["S001", "990101010101", "John Doe", "john@example.com", "CS", 2, 1],  # Valid
            ["S002", None, "Jane Smith", "jane@example.com", "ENG", 3, 2],  # Missing IC
            ["S003", "970303030303", "Bob Johnson", "bob@example.com", "MATH", 1, 1],  # Valid
        ]

        excel_content = self.create_excel_file(headers, data_rows)
        result = parse_excel_data(excel_content, "student")

        # Should only return 2 valid records (S002 is skipped due to missing IC)
        self.assertEqual(len(result), 2)
        self.assertEqual(result[0]["student_id"], "S001")
        self.assertEqual(result[1]["student_id"], "S003")

    def test_parse_invalid_user_type(self):
        """Test parsing with invalid user_type parameter"""
        headers = ["student_id", "ic", "name", "email", "major", "year", "semester"]
        data_rows = [
            ["S001", "990101010101", "John Doe", "john@example.com", "CS", 2, 1],
        ]

        excel_content = self.create_excel_file(headers, data_rows)

        with self.assertRaises(Exception) as context:
            parse_excel_data(excel_content, "invalid_type")

        self.assertIn("Invalid user type", str(context.exception))

    def test_parse_corrupted_excel_file(self):
        """Test parsing a corrupted/invalid Excel file"""
        corrupted_content = b"This is not an Excel file"

        with self.assertRaises(Exception) as context:
            parse_excel_data(corrupted_content, "student")

        self.assertIn("Failed to process Excel file", str(context.exception))


class TestBulkCreateUsers(unittest.TestCase):
    """Test cases for bulk_create_users function"""

    def setUp(self):
        """Set up test data"""
        self.valid_students = [
            {
                "student_id": "S001",
                "ic": "990101010101",
                "name": "John Doe",
                "email": "john@example.com",
                "major": "Computer Science",
                "year": 2,
                "semester": 1,
            },
            {
                "student_id": "S002",
                "ic": "980202020202",
                "name": "Jane Smith",
                "email": "jane@example.com",
                "major": "Engineering",
                "year": 3,
                "semester": 2,
            },
        ]

    # ==================== POSITIVE TEST CASES ====================

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_bulk_create_new_students_success(self, mock_auth, mock_db):
        """Test successful creation of new students"""
        # Mock Firebase Auth - all students are new
        mock_auth.create_user.return_value = None

        # Mock Firestore
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        result = bulk_create_users(self.valid_students, "student")

        self.assertEqual(result["total"], 2)
        self.assertEqual(result["created"], 2)
        self.assertEqual(result["updated"], 0)
        self.assertEqual(result["failed"], 0)
        self.assertEqual(mock_auth.create_user.call_count, 2)
        mock_batch.commit.assert_called_once()

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_bulk_update_existing_students_success(self, mock_auth, mock_db):
        """Test successful update of existing students (password reset)"""
        # Mock Firebase Auth - all students already exist
        mock_auth.create_user.side_effect = exceptions.AlreadyExistsError("User exists")
        mock_auth.update_user.return_value = None

        # Mock Firestore
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        result = bulk_create_users(self.valid_students, "student")

        self.assertEqual(result["total"], 2)
        self.assertEqual(result["created"], 0)
        self.assertEqual(result["updated"], 2)
        self.assertEqual(result["failed"], 0)
        self.assertEqual(mock_auth.update_user.call_count, 2)
        mock_batch.commit.assert_called_once()

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_bulk_create_mixed_new_and_existing(self, mock_auth, mock_db):
        """Test creation with mix of new and existing students"""
        # First student is new, second already exists
        def create_user_side_effect(*args, **kwargs):
            if kwargs.get("uid") == "S001":
                return None  # Success for first student
            else:
                raise exceptions.AlreadyExistsError("User exists")

        mock_auth.create_user.side_effect = create_user_side_effect
        mock_auth.update_user.return_value = None

        # Mock Firestore
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        result = bulk_create_users(self.valid_students, "student")

        self.assertEqual(result["total"], 2)
        self.assertEqual(result["created"], 1)
        self.assertEqual(result["updated"], 1)
        self.assertEqual(result["failed"], 0)

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_bulk_create_password_is_ic_number(self, mock_auth, mock_db):
        """Test that password is set to IC number"""
        mock_auth.create_user.return_value = None

        # Mock Firestore
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        bulk_create_users(self.valid_students, "student")

        # Verify password is set to IC number
        first_call = mock_auth.create_user.call_args_list[0]
        self.assertEqual(first_call[1]["password"], "990101010101")

        second_call = mock_auth.create_user.call_args_list[1]
        self.assertEqual(second_call[1]["password"], "980202020202")

    # ==================== NEGATIVE TEST CASES ====================

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_bulk_create_auth_error_continues_processing(self, mock_auth, mock_db):
        """Test that auth errors don't stop processing of other users"""
        # First student succeeds, second fails, third succeeds
        def create_user_side_effect(*args, **kwargs):
            uid = kwargs.get("uid")
            if uid == "S001":
                return None
            elif uid == "S002":
                raise Exception("Firebase Auth Error")
            else:
                return None

        mock_auth.create_user.side_effect = create_user_side_effect

        # Mock Firestore
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        # Add a third student
        students = self.valid_students + [
            {
                "student_id": "S003",
                "ic": "970303030303",
                "name": "Bob Johnson",
                "email": "bob@example.com",
                "major": "Mathematics",
                "year": 1,
                "semester": 1,
            }
        ]

        result = bulk_create_users(students, "student")

        self.assertEqual(result["total"], 3)
        self.assertEqual(result["created"], 2)
        self.assertEqual(result["failed"], 1)
        self.assertGreater(len(result["errors"]), 0)
        self.assertIn("Auth Error", result["errors"][0])

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_bulk_create_firestore_batch_error(self, mock_auth, mock_db):
        """Test handling of Firestore batch commit error"""
        mock_auth.create_user.return_value = None

        # Mock Firestore with batch commit error
        mock_batch = MagicMock()
        mock_batch.commit.side_effect = Exception("Firestore Batch Error: Connection timeout")
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        result = bulk_create_users(self.valid_students, "student")

        # Auth operations should succeed, but batch commit fails
        self.assertEqual(result["created"], 2)
        self.assertIn("Firestore Batch Error", result["errors"][0])

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_bulk_create_invalid_email_format(self, mock_auth, mock_db):
        """Test handling of invalid email format in Firebase Auth"""
        mock_auth.create_user.side_effect = Exception("Invalid email format")

        # Mock Firestore
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        students = [
            {
                "student_id": "S001",
                "ic": "990101010101",
                "name": "John Doe",
                "email": "invalid-email",  # Invalid email
                "major": "Computer Science",
                "year": 2,
                "semester": 1,
            }
        ]

        result = bulk_create_users(students, "student")

        self.assertEqual(result["failed"], 1)
        self.assertIn("Auth Error", result["errors"][0])
        self.assertIn("invalid-email", result["errors"][0])

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_bulk_create_empty_user_list(self, mock_auth, mock_db):
        """Test bulk create with empty user list"""
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch

        result = bulk_create_users([], "student")

        self.assertEqual(result["total"], 0)
        self.assertEqual(result["created"], 0)
        self.assertEqual(result["updated"], 0)
        self.assertEqual(result["failed"], 0)

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_bulk_create_lecturer_instead_of_student(self, mock_auth, mock_db):
        """Test bulk create with lecturer type instead of student"""
        mock_auth.create_user.return_value = None

        # Mock Firestore
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        lecturers = [
            {
                "lecturer_id": "L001",
                "ic": "990101010101",
                "name": "Dr. Smith",
                "email": "smith@example.com",
                "faculty": "Engineering",
            }
        ]

        result = bulk_create_users(lecturers, "lecturer")

        self.assertEqual(result["created"], 1)
        # Verify lecturer_id field is used instead of student_id
        call_args = mock_auth.create_user.call_args[1]
        self.assertEqual(call_args["uid"], "L001")


class TestIntegrationScenarios(unittest.TestCase):
    """Integration test scenarios combining parse and bulk create"""

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_full_import_workflow_success(self, mock_auth, mock_db):
        """Test complete workflow: parse Excel -> bulk create"""
        # Create Excel file
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ["student_id", "ic", "name", "email", "major", "year", "semester"]
        sheet.append(headers)
        sheet.append(["S001", "990101010101", "John Doe", "john@example.com", "CS", 2, 1])

        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        excel_content = excel_buffer.read()

        # Parse Excel
        parsed_users = parse_excel_data(excel_content, "student")
        self.assertEqual(len(parsed_users), 1)

        # Mock Firebase for bulk create
        mock_auth.create_user.return_value = None
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        # Bulk create
        result = bulk_create_users(parsed_users, "student")

        self.assertEqual(result["created"], 1)
        self.assertEqual(result["failed"], 0)

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_full_import_workflow_with_errors(self, mock_auth, mock_db):
        """Test workflow with some records failing"""
        # Create Excel with mixed valid/invalid data
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ["student_id", "ic", "name", "email", "major", "year", "semester"]
        sheet.append(headers)
        sheet.append(["S001", "990101010101", "John Doe", "john@example.com", "CS", 2, 1])
        sheet.append(["S002", None, "Jane Smith", "jane@example.com", "ENG", 3, 2])  # Missing IC

        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        excel_content = excel_buffer.read()

        # Parse Excel - should skip invalid row
        parsed_users = parse_excel_data(excel_content, "student")
        self.assertEqual(len(parsed_users), 1)  # Only 1 valid record

        # Mock Firebase
        mock_auth.create_user.return_value = None
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        # Bulk create
        result = bulk_create_users(parsed_users, "student")

        self.assertEqual(result["created"], 1)
        self.assertEqual(result["failed"], 0)


if __name__ == "__main__":
    unittest.main()