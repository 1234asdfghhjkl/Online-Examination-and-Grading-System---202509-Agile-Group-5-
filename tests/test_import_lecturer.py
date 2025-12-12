# tests/test_import_lecturer.py
import unittest
from unittest.mock import patch, MagicMock
from io import BytesIO
import openpyxl
from firebase_admin import exceptions

# Import the functions to test
from services.user_service import parse_excel_data, bulk_create_users


class TestParseExcelDataLecturer(unittest.TestCase):
    """Test cases for parse_excel_data function with lecturer type"""

    def create_excel_file(self, headers, data_rows):
        """Helper function to create an in-memory Excel file"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)

        # Write data rows
        for row_idx, row_data in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        # Save to BytesIO
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.read()

    # ==================== POSITIVE TEST CASES ====================

    def test_parse_valid_lecturer_excel(self):
        """Test parsing a valid lecturer Excel file with all required fields"""
        headers = ["lecturer_id", "ic", "name", "email", "faculty"]
        data_rows = [
            ["L001", "800101010101", "Dr. John Smith", "john.smith@university.edu", "Engineering"],
            ["L002", "750202020202", "Prof. Jane Doe", "jane.doe@university.edu", "Computer Science"],
        ]

        excel_content = self.create_excel_file(headers, data_rows)
        result = parse_excel_data(excel_content, "lecturer")

        self.assertEqual(len(result), 2)
        self.assertEqual(result[0]["lecturer_id"], "L001")
        self.assertEqual(result[0]["name"], "Dr. John Smith")
        self.assertEqual(result[0]["faculty"], "Engineering")
        self.assertEqual(result[1]["lecturer_id"], "L002")
        self.assertEqual(result[1]["faculty"], "Computer Science")

    def test_parse_lecturer_excel_with_extra_columns(self):
        """Test parsing Excel with extra columns (should be ignored)"""
        headers = ["lecturer_id", "ic", "name", "email", "faculty", "phone", "office_number"]
        data_rows = [
            ["L001", "800101010101", "Dr. Smith", "smith@university.edu", "Engineering", "555-1234", "A-123"],
        ]

        excel_content = self.create_excel_file(headers, data_rows)
        result = parse_excel_data(excel_content, "lecturer")

        self.assertEqual(len(result), 1)
        self.assertNotIn("phone", result[0])
        self.assertNotIn("office_number", result[0])
        self.assertIn("lecturer_id", result[0])
        self.assertIn("faculty", result[0])

    def test_parse_lecturer_excel_mixed_case_headers(self):
        """Test parsing Excel with mixed case headers (should work with .lower())"""
        headers = ["Lecturer_ID", "IC", "Name", "EMAIL", "Faculty"]
        data_rows = [
            ["L001", "800101010101", "Dr. Smith", "smith@university.edu", "Engineering"],
        ]

        excel_content = self.create_excel_file(headers, data_rows)
        result = parse_excel_data(excel_content, "lecturer")

        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["lecturer_id"], "L001")
        self.assertEqual(result[0]["faculty"], "Engineering")

    def test_parse_lecturer_excel_skip_empty_rows(self):
        """Test that empty rows are skipped"""
        headers = ["lecturer_id", "ic", "name", "email", "faculty"]
        data_rows = [
            ["L001", "800101010101", "Dr. Smith", "smith@university.edu", "Engineering"],
            [None, None, None, None, None],  # Empty row
            ["L002", "750202020202", "Prof. Doe", "doe@university.edu", "Science"],
        ]

        excel_content = self.create_excel_file(headers, data_rows)
        result = parse_excel_data(excel_content, "lecturer")

        self.assertEqual(len(result), 2)
        self.assertEqual(result[0]["lecturer_id"], "L001")
        self.assertEqual(result[1]["lecturer_id"], "L002")

    def test_parse_lecturer_excel_with_whitespace(self):
        """Test that whitespace in IDs and emails is properly stripped"""
        headers = ["lecturer_id", "ic", "name", "email", "faculty"]
        data_rows = [
            [" L001 ", " 800101010101 ", "Dr. Smith", " smith@university.edu ", "Engineering"],
        ]

        excel_content = self.create_excel_file(headers, data_rows)
        result = parse_excel_data(excel_content, "lecturer")

        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["lecturer_id"], "L001")
        self.assertEqual(result[0]["ic"], "800101010101")
        self.assertEqual(result[0]["email"], "smith@university.edu")

    def test_parse_lecturer_excel_numeric_id(self):
        """Test that numeric lecturer IDs are converted to strings"""
        headers = ["lecturer_id", "ic", "name", "email", "faculty"]
        data_rows = [
            [1001, 800101010101, "Dr. Smith", "smith@university.edu", "Engineering"],
        ]

        excel_content = self.create_excel_file(headers, data_rows)
        result = parse_excel_data(excel_content, "lecturer")

        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["lecturer_id"], "1001")
        self.assertEqual(result[0]["ic"], "800101010101")

    # ==================== NEGATIVE TEST CASES ====================

    def test_parse_lecturer_excel_missing_required_column(self):
        """Test parsing Excel missing a required column (should raise Exception)"""
        headers = ["lecturer_id", "ic", "name", "email"]  # Missing 'faculty'
        data_rows = [
            ["L001", "800101010101", "Dr. Smith", "smith@university.edu"],
        ]

        excel_content = self.create_excel_file(headers, data_rows)

        with self.assertRaises(Exception) as context:
            parse_excel_data(excel_content, "lecturer")

        self.assertIn("Failed to process Excel file", str(context.exception))
        self.assertIn("faculty", str(context.exception).lower())

    def test_parse_lecturer_excel_no_valid_records(self):
        """Test parsing Excel with no valid records (should raise Exception)"""
        headers = ["lecturer_id", "ic", "name", "email", "faculty"]
        data_rows = [
            [None, None, None, None, None],  # All empty
            ["", "", "", "", ""],  # All blank strings
        ]

        excel_content = self.create_excel_file(headers, data_rows)

        with self.assertRaises(Exception) as context:
            parse_excel_data(excel_content, "lecturer")

        self.assertIn("No valid user records found", str(context.exception))

    def test_parse_lecturer_excel_incomplete_data(self):
        """Test parsing Excel with incomplete data in rows (should skip incomplete rows)"""
        headers = ["lecturer_id", "ic", "name", "email", "faculty"]
        data_rows = [
            ["L001", "800101010101", "Dr. Smith", "smith@university.edu", "Engineering"],  # Valid
            ["L002", None, "Prof. Doe", "doe@university.edu", "Science"],  # Missing IC
            ["L003", "770303030303", "Dr. Johnson", "johnson@university.edu", "Mathematics"],  # Valid
        ]

        excel_content = self.create_excel_file(headers, data_rows)
        result = parse_excel_data(excel_content, "lecturer")

        # Should only return 2 valid records (L002 is skipped due to missing IC)
        self.assertEqual(len(result), 2)
        self.assertEqual(result[0]["lecturer_id"], "L001")
        self.assertEqual(result[1]["lecturer_id"], "L003")

    def test_parse_lecturer_excel_missing_email(self):
        """Test parsing Excel with missing email field"""
        headers = ["lecturer_id", "ic", "name", "email", "faculty"]
        data_rows = [
            ["L001", "800101010101", "Dr. Smith", "smith@university.edu", "Engineering"],  # Valid
            ["L002", "750202020202", "Prof. Doe", None, "Science"],  # Missing email
        ]

        excel_content = self.create_excel_file(headers, data_rows)
        result = parse_excel_data(excel_content, "lecturer")

        # Should only return 1 valid record
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["lecturer_id"], "L001")

    def test_parse_lecturer_excel_missing_name(self):
        """Test parsing Excel with missing name field"""
        headers = ["lecturer_id", "ic", "name", "email", "faculty"]
        data_rows = [
            ["L001", "800101010101", "Dr. Smith", "smith@university.edu", "Engineering"],  # Valid
            ["L002", "750202020202", None, "doe@university.edu", "Science"],  # Missing name
        ]

        excel_content = self.create_excel_file(headers, data_rows)
        result = parse_excel_data(excel_content, "lecturer")

        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["lecturer_id"], "L001")


class TestBulkCreateUsersLecturer(unittest.TestCase):
    """Test cases for bulk_create_users function with lecturer type"""

    def setUp(self):
        """Set up test data"""
        self.valid_lecturers = [
            {
                "lecturer_id": "L001",
                "ic": "800101010101",
                "name": "Dr. John Smith",
                "email": "john.smith@university.edu",
                "faculty": "Engineering",
            },
            {
                "lecturer_id": "L002",
                "ic": "750202020202",
                "name": "Prof. Jane Doe",
                "email": "jane.doe@university.edu",
                "faculty": "Computer Science",
            },
        ]

    # ==================== POSITIVE TEST CASES ====================

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_bulk_create_new_lecturers_success(self, mock_auth, mock_db):
        """Test successful creation of new lecturers"""
        # Mock Firebase Auth - all lecturers are new
        mock_auth.create_user.return_value = None

        # Mock Firestore
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        result = bulk_create_users(self.valid_lecturers, "lecturer")

        self.assertEqual(result["total"], 2)
        self.assertEqual(result["created"], 2)
        self.assertEqual(result["updated"], 0)
        self.assertEqual(result["failed"], 0)
        self.assertEqual(mock_auth.create_user.call_count, 2)
        mock_batch.commit.assert_called_once()

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_bulk_update_existing_lecturers_success(self, mock_auth, mock_db):
        """Test successful update of existing lecturers (password reset)"""
        # Mock Firebase Auth - all lecturers already exist
        mock_auth.create_user.side_effect = exceptions.AlreadyExistsError("User exists")
        mock_auth.update_user.return_value = None

        # Mock Firestore
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        result = bulk_create_users(self.valid_lecturers, "lecturer")

        self.assertEqual(result["total"], 2)
        self.assertEqual(result["created"], 0)
        self.assertEqual(result["updated"], 2)
        self.assertEqual(result["failed"], 0)
        self.assertEqual(mock_auth.update_user.call_count, 2)
        mock_batch.commit.assert_called_once()

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_bulk_create_mixed_new_and_existing(self, mock_auth, mock_db):
        """Test creation with mix of new and existing lecturers"""
        # First lecturer is new, second already exists
        def create_user_side_effect(*args, **kwargs):
            if kwargs.get("uid") == "L001":
                return None  # Success for first lecturer
            else:
                raise exceptions.AlreadyExistsError("User exists")

        mock_auth.create_user.side_effect = create_user_side_effect
        mock_auth.update_user.return_value = None

        # Mock Firestore
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        result = bulk_create_users(self.valid_lecturers, "lecturer")

        self.assertEqual(result["total"], 2)
        self.assertEqual(result["created"], 1)
        self.assertEqual(result["updated"], 1)
        self.assertEqual(result["failed"], 0)

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_bulk_create_password_is_ic_number(self, mock_auth, mock_db):
        """Test that password is set to IC number"""
        mock_auth.create_user.return_value = None

        # Mock Firestore
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        bulk_create_users(self.valid_lecturers, "lecturer")

        # Verify password is set to IC number
        first_call = mock_auth.create_user.call_args_list[0]
        self.assertEqual(first_call[1]["password"], "800101010101")

        second_call = mock_auth.create_user.call_args_list[1]
        self.assertEqual(second_call[1]["password"], "750202020202")

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_bulk_create_lecturer_profile_data(self, mock_auth, mock_db):
        """Test that lecturer profile data includes faculty field"""
        mock_auth.create_user.return_value = None

        # Mock Firestore
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_collection = MagicMock()
        mock_db.collection.return_value = mock_collection

        bulk_create_users(self.valid_lecturers, "lecturer")

        # Verify Firestore batch.set was called with correct data
        set_calls = mock_batch.set.call_args_list
        self.assertEqual(len(set_calls), 2)

        # Check first lecturer's profile data
        first_profile = set_calls[0][0][1]  # Second argument to set()
        self.assertEqual(first_profile["role"], "lecturer")
        self.assertEqual(first_profile["faculty"], "Engineering")
        self.assertEqual(first_profile["lecturer_id"], "L001")
        self.assertIn("uid", first_profile)
        self.assertIn("email", first_profile)
        self.assertIn("name", first_profile)
        self.assertIn("ic", first_profile)

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_bulk_create_uses_lecturer_id_as_uid(self, mock_auth, mock_db):
        """Test that lecturer_id is used as the Firebase uid"""
        mock_auth.create_user.return_value = None

        # Mock Firestore
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        bulk_create_users(self.valid_lecturers, "lecturer")

        # Verify uid matches lecturer_id
        first_call = mock_auth.create_user.call_args_list[0]
        self.assertEqual(first_call[1]["uid"], "L001")

        second_call = mock_auth.create_user.call_args_list[1]
        self.assertEqual(second_call[1]["uid"], "L002")

    # ==================== NEGATIVE TEST CASES ====================

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_bulk_create_auth_error_continues_processing(self, mock_auth, mock_db):
        """Test that auth errors don't stop processing of other lecturers"""
        # First lecturer succeeds, second fails, third succeeds
        def create_user_side_effect(*args, **kwargs):
            uid = kwargs.get("uid")
            if uid == "L001":
                return None
            elif uid == "L002":
                raise Exception("Firebase Auth Error")
            else:
                return None

        mock_auth.create_user.side_effect = create_user_side_effect

        # Mock Firestore
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        # Add a third lecturer
        lecturers = self.valid_lecturers + [
            {
                "lecturer_id": "L003",
                "ic": "700303030303",
                "name": "Dr. Bob Johnson",
                "email": "bob.johnson@university.edu",
                "faculty": "Mathematics",
            }
        ]

        result = bulk_create_users(lecturers, "lecturer")

        self.assertEqual(result["total"], 3)
        self.assertEqual(result["created"], 2)
        self.assertEqual(result["failed"], 1)
        self.assertGreater(len(result["errors"]), 0)
        self.assertIn("Auth Error", result["errors"][0])

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_bulk_create_firestore_batch_error(self, mock_auth, mock_db):
        """Test handling of Firestore batch commit error"""
        mock_auth.create_user.return_value = None

        # Mock Firestore with batch commit error
        mock_batch = MagicMock()
        mock_batch.commit.side_effect = Exception("Firestore Batch Error: Connection timeout")
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        result = bulk_create_users(self.valid_lecturers, "lecturer")

        # Auth operations should succeed, but batch commit fails
        self.assertEqual(result["created"], 2)
        self.assertIn("Firestore Batch Error", result["errors"][0])

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_bulk_create_invalid_email_format(self, mock_auth, mock_db):
        """Test handling of invalid email format in Firebase Auth"""
        mock_auth.create_user.side_effect = Exception("Invalid email format")

        # Mock Firestore
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        lecturers = [
            {
                "lecturer_id": "L001",
                "ic": "800101010101",
                "name": "Dr. Smith",
                "email": "invalid-email",  # Invalid email
                "faculty": "Engineering",
            }
        ]

        result = bulk_create_users(lecturers, "lecturer")

        self.assertEqual(result["failed"], 1)
        self.assertIn("Auth Error", result["errors"][0])
        self.assertIn("invalid-email", result["errors"][0])

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_bulk_create_empty_lecturer_list(self, mock_auth, mock_db):
        """Test bulk create with empty lecturer list"""
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch

        result = bulk_create_users([], "lecturer")

        self.assertEqual(result["total"], 0)
        self.assertEqual(result["created"], 0)
        self.assertEqual(result["updated"], 0)
        self.assertEqual(result["failed"], 0)

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_bulk_create_duplicate_email_error(self, mock_auth, mock_db):
        """Test handling of duplicate email addresses"""
        mock_auth.create_user.side_effect = Exception("Email already exists")

        # Mock Firestore
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        result = bulk_create_users(self.valid_lecturers, "lecturer")

        self.assertEqual(result["failed"], 2)
        self.assertEqual(len(result["errors"]), 2)

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_bulk_create_weak_password_error(self, mock_auth, mock_db):
        """Test handling of weak password (IC too short)"""
        mock_auth.create_user.side_effect = Exception("Password must be at least 6 characters")

        # Mock Firestore
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        lecturers = [
            {
                "lecturer_id": "L001",
                "ic": "12345",  # Too short IC
                "name": "Dr. Smith",
                "email": "smith@university.edu",
                "faculty": "Engineering",
            }
        ]

        result = bulk_create_users(lecturers, "lecturer")

        self.assertEqual(result["failed"], 1)
        self.assertIn("Auth Error", result["errors"][0])


class TestIntegrationScenariosLecturer(unittest.TestCase):
    """Integration test scenarios combining parse and bulk create for lecturers"""

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_full_import_workflow_success(self, mock_auth, mock_db):
        """Test complete workflow: parse Excel -> bulk create"""
        # Create Excel file
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ["lecturer_id", "ic", "name", "email", "faculty"]
        sheet.append(headers)
        sheet.append(["L001", "800101010101", "Dr. Smith", "smith@university.edu", "Engineering"])

        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        excel_content = excel_buffer.read()

        # Parse Excel
        parsed_users = parse_excel_data(excel_content, "lecturer")
        self.assertEqual(len(parsed_users), 1)
        self.assertEqual(parsed_users[0]["faculty"], "Engineering")

        # Mock Firebase for bulk create
        mock_auth.create_user.return_value = None
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        # Bulk create
        result = bulk_create_users(parsed_users, "lecturer")

        self.assertEqual(result["created"], 1)
        self.assertEqual(result["failed"], 0)

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_full_import_workflow_with_errors(self, mock_auth, mock_db):
        """Test workflow with some records failing"""
        # Create Excel with mixed valid/invalid data
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ["lecturer_id", "ic", "name", "email", "faculty"]
        sheet.append(headers)
        sheet.append(["L001", "800101010101", "Dr. Smith", "smith@university.edu", "Engineering"])
        sheet.append(["L002", None, "Prof. Doe", "doe@university.edu", "Science"])  # Missing IC

        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        excel_content = excel_buffer.read()

        # Parse Excel - should skip invalid row
        parsed_users = parse_excel_data(excel_content, "lecturer")
        self.assertEqual(len(parsed_users), 1)  # Only 1 valid record

        # Mock Firebase
        mock_auth.create_user.return_value = None
        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        # Bulk create
        result = bulk_create_users(parsed_users, "lecturer")

        self.assertEqual(result["created"], 1)
        self.assertEqual(result["failed"], 0)

    @patch("services.user_service.db")
    @patch("services.user_service.auth")
    def test_full_import_workflow_multiple_lecturers(self, mock_auth, mock_db):
        """Test workflow with multiple valid lecturers"""
        # Create Excel with multiple lecturers
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ["lecturer_id", "ic", "name", "email", "faculty"]
        sheet.append(headers)
        sheet.append(["L001", "800101010101", "Dr. Smith", "smith@university.edu", "Engineering"])
        sheet.append(["L002", "750202020202", "Prof. Doe", "doe@university.edu", "Computer Science"])
        sheet.append(["L003", "700303030303", "Dr. Johnson", "johnson@university.edu", "Mathematics"])

        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        excel_content = excel_buffer.read()

        # Parse Excel
        parsed_users = parse_excel_data(excel_content, "lecturer")
        self.assertEqual(len(parsed_users), 3)

        # Mock Firebase - first two are new, third already exists
        def create_user_side_effect(*args, **kwargs):
            if kwargs.get("uid") in ["L001", "L002"]:
                return None
            else:
                raise exceptions.AlreadyExistsError("User exists")

        mock_auth.create_user.side_effect = create_user_side_effect
        mock_auth.update_user.return_value = None

        mock_batch = MagicMock()
        mock_db.batch.return_value = mock_batch
        mock_db.collection.return_value.document.return_value = MagicMock()

        # Bulk create
        result = bulk_create_users(parsed_users, "lecturer")

        self.assertEqual(result["total"], 3)
        self.assertEqual(result["created"], 2)
        self.assertEqual(result["updated"], 1)
        self.assertEqual(result["failed"], 0)


if __name__ == "__main__":
    unittest.main()